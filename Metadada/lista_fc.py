import arcpy
import os
import openpyxl

gdb = r""
output = r""


wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Camadas"

linha = 2

arcpy.env.workspace = gdb
for fd in arcpy.ListDatasets():
    print(fd)
    arcpy.env.workspace = os.path.join(gdb, fd)
    fcs = arcpy.ListFeatureClasses()
    for fc in fcs:
        print(fc)
        c1 = ws1.cell(linha, 1)
        c1.value = fd
        c2 = ws1.cell(linha, 2)
        c2.value = fc
        linha += 1

wb.save(output)
