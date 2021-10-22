
import xml.etree.ElementTree as ET
import openpyxl
import pandas

xml_base = r""
excel_branco = r""
excel_base = r""
df = pandas.read_excel(excel_base)
titles_base = df['NOME DA CAMADA'].tolist()

tree_1 = ET.parse(xml_base)
root = tree_1.getroot()
layers = root.findall("./{http://www.opengis.net/wms}Capability/{http://www.opengis.net/wms}Layer/{http://www.opengis.net/wms}Layer")
linha = 1
wb_write = openpyxl.Workbook()
sheet_write = wb_write.active

for child in layers:
    title = child.find("{http://www.opengis.net/wms}Title").text

    if title not in titles_base:
         abstract = child.find("{http://www.opengis.net/wms}Abstract").text
         print(title, abstract)
         c1 = sheet_write.cell(linha, 2)
         c1.value = title
         c2 = sheet_write.cell(linha, 7)
         c2.value = abstract
         linha += 1
wb_write.save(excel_branco)


