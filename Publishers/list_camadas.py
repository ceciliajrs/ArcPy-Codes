# -*- coding: utf-8 -*-

import arcpy
import openpyxl
import os

pasta = r""
excel_branco = r""

linha = 2
wb = openpyxl.Workbook()
ws1 = wb.active

for root, dirs, files in os.walk(pasta):
    # print "Root: ", root
    # print "Dirs: ", dirs
    # print "Files: ", files
    for pasta_mxd in dirs:
        print "PASTA:", pasta_mxd
        caminho = ""
        if pasta_mxd == "Tomo_I":
            caminho = os.path.join(pasta, "", "")
        elif pasta_mxd == "Tomo_II":
            caminho = os.path.join(pasta, "", "")
        elif pasta_mxd == "Apendice_1":
            caminho = os.path.join(pasta, "", "", "")
        else:
            caminho = os.path.join(pasta, pasta_mxd)
        arcpy.env.workspace = caminho
        mxd_list = arcpy.ListFiles("*.mxd")
        for mxd in mxd_list:
            mxd_active = arcpy.mapping.MapDocument(os.path.join(caminho, mxd))
            camadas_list = arcpy.mapping.ListLayers(mxd_active)
            for camada in camadas_list:
                if camada.supports("DATASOURCE"):
                    print pasta_mxd, mxd, camada.name
                    c0 = ws1.cell(linha, 1)
                    c0.value = pasta_mxd
                    c1 = ws1.cell(linha, 2)
                    c1.value = mxd
                    c2 = ws1.cell(linha, 3)
                    c2.value = camada.name
                    c3 = ws1.cell(linha, 4)
                    c3.value = camada.dataSource
                    c0 = ws1.cell(1, 1)
                    c0.value = "PASTA"
                    c1 = ws1.cell(1, 2)
                    c1.value = "MXD"
                    c2 = ws1.cell(1, 3)
                    c2.value = "CAMADA"
                    c3 = ws1.cell(1, 4)
                    c3.value = "DATA SOURCE"
                    linha += 1
        wb.save(excel_branco)

