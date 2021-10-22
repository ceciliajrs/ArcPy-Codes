import arcpy
import os
import openpyxl

GDB = r"..."

output = r"..."


wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Camadas"

linha = 2
arcpy.env.workspace = GDB
featuredatasets = arcpy.ListDatasets()
for dataset in featuredatasets:
    print("Dataset: ", dataset)
    arcpy.env.workspace = os.path.join(GDB, dataset)
    featureclasses = arcpy.ListFeatureClasses()
    for fc in featureclasses:
        print("Feature Class: ", fc)
        arcpy.env.workspace = os.path.join(GDB, dataset, fc)
        fields = arcpy.ListFields(fc)
        for field in fields:
            print(field.name)
            c1 = ws1.cell(linha, 1)
            c1.value = dataset
            c2 = ws1.cell(linha, 2)
            c2.value = fc
            c3 = ws1.cell(linha, 3)
            c3.value = field.name
            c4 = ws1.cell(linha, 4)
            c4.value = field.type
            linha += 1
wb.save(output)
print(linha)

